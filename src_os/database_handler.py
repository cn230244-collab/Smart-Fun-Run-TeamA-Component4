import os
import openpyxl
from openpyxl.utils import get_column_letter

# Shared Resource: The Excel file where live data is stored and retrieved [3]
EXCEL_FILE = 'hydration_log.xlsx'

def inisialisasi_excel():
    """
    OS Mechanism: Automated File Initialization.
    Ensures the persistence of data by checking for the file's existence 
    at system startup. If not found, it creates the database with 
    standardized OS tracking headers. [3]
    """
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Log Sistem OS-IoT"
        
        # Headers include OS metrics: PID and Thread ID for system monitoring [3]
        headers = ["Tarikh & Masa", "Process ID", "Thread ID", "IP Address", "Quantity (Botol)", "Status"]
        ws.append(headers)
        wb.save(EXCEL_FILE)
        print(f"🎉 System initialization complete: {EXCEL_FILE} created successfully.")

def simpan_ke_excel(data_list):
    """
    OS Mechanism: File Management & Data Persistence.
    Appends live sensor data into the shared Excel resource. 
    Includes a dynamic formatting algorithm to ensure readability. [2, 3]
    """
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Append the new data row sent from the multithreaded server
    ws.append(data_list)
    
    # DYNAMIC FORMATTING LOGIC: 
    # Automatically calculates the maximum character length of each column.
    # Adjusts column dimensions in real-time to prevent data overlapping.
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col.column)
        for cell in col:
            if cell.value:
                # Calculate required width based on data length
                max_len = max(max_len, len(str(cell.value)))
        
        # Apply the 'Auto-Fit' width to the file system
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    wb.save(EXCEL_FILE)
