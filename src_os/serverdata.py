import os
from flask import Flask, request, jsonify
import threading
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# Mutex Lock untuk memastikan integriti data (Concurrency Control)
db_lock = threading.Lock()

# 🎯 NAMA FAIL BARU DI SINI (Ditukar kepada hydration_log.xlsx)
EXCEL_FILE = 'hydration_log.xlsx'

def inisialisasi_excel():
    """Fungsi untuk mencipta fail Excel bertema profesional dengan kolum auto-lebar"""
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Log Sistem OS-IoT"
        ws.views.sheetView[0].showGridLines = True
        
        # Tetapan Gaya Papan Atas (Header) - Tema Biru Korporat (Corporate Blue)
        headers = ["Tarikh & Masa", "Process ID", "Thread ID", "IP Address", "Quantity (Botol)", "Status"]
        font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        align_center = Alignment(horizontal="center", vertical="center")
        
        ws.append(headers)
        
        # Aplikasi gaya pada header
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            
        wb.save(EXCEL_FILE)
        print(f"🎉 Fail {EXCEL_FILE} baharu dengan format auto-lebar berjaya dicipta!")

@app.route('/api/funrun', methods=['POST'])
def receive_bottle_data():
    data = request.get_json()
    
    # 1. Mengekstrak data IoT (API Contract)
    quantity = data.get('quantity')
    status_botol = data.get('status')
    
    # 2. Menangkap Metrik Sistem Operasi (OS & Network)
    waktu_sekarang = datetime.now().strftime('%d/%m/%Y %H:%M')
    thread_id = threading.current_thread().name
    process_id = os.getpid()
    ip_address = request.remote_addr

    # 3. Mutex Lock - Menghalang Race Condition pada fail Excel
    with db_lock:
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            
            # Gaya Teks Badan Data
            font_body = Font(name="Arial", size=11)
            fill_zebra = PatternFill(start_color="F2F6F9", end_color="F2F6F9", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9')
            )
            
            # Tambah baris data baharu
            row_data = [waktu_sekarang, process_id, thread_id, ip_address, quantity, status_botol]
            ws.append(row_data)
            
            row_idx = ws.max_row
            
            # Kemasan Penyusunan Lajur & Corak Zebra Striping
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = font_body
                cell.border = thin_border
                
                # Aplikasi warna selang-seli (Zebra) untuk baris genap
                if row_idx % 2 == 0:
                    cell.fill = fill_zebra
                
                # Format susunan teks (Alignment)
                if col_idx in [1, 3, 4]:  # Masa, Thread, IP
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_idx in [2, 5]:   # PID, Kuantiti
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = "#,##0"
                else:                     # Status
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Auto-Fit Column Width (Automatik Kasi Lebar!)
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
                
            wb.save(EXCEL_FILE)
            print(f"📡 [THREAD {thread_id}] Sukses merekod -> Qty: {quantity}, IP: {ip_address}")
            return jsonify({"message": "Data berjaya direkod ke Excel", "status": "success"}), 200
            
        except Exception as e:
            print(f"❌ Ralat sistem fail: {str(e)}")
            return jsonify({"message": "Gagal merekod data", "status": "error"}), 500

if __name__ == '__main__':
    # Memastikan fail Excel dicipta siap-siap sebelum menerima data
    inisialisasi_excel()
    
    # Wajib host='0.0.0.0' supaya cip IoT kawan anda boleh nampak server laptop anda
    app.run(debug=True, host='0.0.0.0', port=5000)